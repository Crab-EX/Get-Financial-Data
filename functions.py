import io
import re
import unicodedata
import zipfile
from typing import Dict

import openpyxl


MINIMAL_STYLES_XML = b'''<?xml version="1.0" encoding="UTF-8" standalone="yes"?>
<styleSheet xmlns="http://schemas.openxmlformats.org/spreadsheetml/2006/main">
    <numFmts count="0"/>
    <fonts count="1">
        <font>
            <sz val="11"/>
            <color theme="1"/>
            <name val="Calibri"/>
            <family val="2"/>
            <scheme val="minor"/>
        </font>
    </fonts>
    <fills count="2">
        <fill><patternFill patternType="none"/></fill>
        <fill><patternFill patternType="gray125"/></fill>
    </fills>
    <borders count="1">
        <border><left/><right/><top/><bottom/><diagonal/></border>
    </borders>
    <cellStyleXfs count="1">
        <xf numFmtId="0" fontId="0" fillId="0" borderId="0"/>
    </cellStyleXfs>
    <cellXfs count="1">
        <xf numFmtId="0" fontId="0" fillId="0" borderId="0" xfId="0"/>
    </cellXfs>
    <cellStyles count="1">
        <cellStyle name="Normal" xfId="0" builtinId="0"/>
    </cellStyles>
</styleSheet>
'''


def _normalize_cell_value(value) -> str:
    """Normalize cell values so keys can be compared safely."""
    if isinstance(value, float) and value.is_integer():
        value = int(value)
    return str(value).strip() if value is not None else ""


def _normalize_text(value) -> str:
    """Normalize text for robust matching (accent/case/spacing insensitive)."""
    text = _normalize_cell_value(value)
    text = unicodedata.normalize("NFKD", text)
    text = "".join(ch for ch in text if not unicodedata.combining(ch))
    text = re.sub(r"\s+", " ", text)
    return text.casefold().strip()


def _resolve_sheet_name(workbook, requested_sheet_name: str):
    """Return an exact or normalized-matching sheet name from workbook."""
    if requested_sheet_name in workbook.sheetnames:
        return requested_sheet_name

    normalized_requested = _normalize_text(requested_sheet_name)
    for workbook_sheet_name in workbook.sheetnames:
        if _normalize_text(workbook_sheet_name) == normalized_requested:
            return workbook_sheet_name
    return None


def _get_uploaded_file_payload(
    uploaded_files_by_name: Dict[str, bytes],
    registration_number: str,
    year: str,
):
    """Support both underscore and dash naming conventions."""
    registration_number = registration_number.strip()
    year = year.strip()
    candidate_names = [
        f"{registration_number}_{year}.xlsx",
        f"{registration_number}-{year}.xlsx",
    ]
    for candidate_name in candidate_names:
        if candidate_name in uploaded_files_by_name:
            return candidate_name, uploaded_files_by_name[candidate_name]
    return None, None


def _load_workbook_safe(workbook_bytes: bytes, data_only: bool = True):
    """
    Load workbook with fallbacks for malformed style metadata.

    Some third-party exports include invalid fill/style definitions that can trigger:
    "expected <class 'openpyxl.styles.fills.Fill'>".
    """
    load_attempts = [
        {"data_only": data_only, "read_only": False},
        {"data_only": data_only, "read_only": True},
        {"data_only": False, "read_only": True},
    ]

    def _try_load(bytes_payload: bytes):
        last_ex = None
        for attempt in load_attempts:
            try:
                workbook = openpyxl.load_workbook(io.BytesIO(bytes_payload), **attempt)
                return workbook, None
            except Exception as ex:
                last_ex = ex
        return None, last_ex

    last_error = None
    workbook, last_error = _try_load(workbook_bytes)
    if workbook is not None:
        return workbook

    # Fallback: replace malformed style sheet and retry.
    repaired_payload = _sanitize_xlsx_styles(workbook_bytes)
    if repaired_payload is not None:
        workbook, repaired_error = _try_load(repaired_payload)
        if workbook is not None:
            return workbook
        last_error = repaired_error

    raise ValueError(
        "Unable to read XLSX file. The file seems to contain unsupported/corrupted style definitions. "
        "Please re-save it as a standard .xlsx file (e.g., open and Save As in Excel) and try again. "
        f"Original error: {last_error}"
    )


def _sanitize_xlsx_styles(workbook_bytes: bytes):
    """Return a copy of XLSX bytes with a minimal valid styles.xml, or None if not a valid zip."""
    input_buffer = io.BytesIO(workbook_bytes)
    output_buffer = io.BytesIO()

    try:
        with zipfile.ZipFile(input_buffer, "r") as input_zip:
            with zipfile.ZipFile(output_buffer, "w", compression=zipfile.ZIP_DEFLATED) as output_zip:
                found_styles = False
                for item in input_zip.infolist():
                    item_name = item.filename
                    if item_name == "xl/styles.xml":
                        found_styles = True
                        output_zip.writestr(item_name, MINIMAL_STYLES_XML)
                        continue
                    output_zip.writestr(item_name, input_zip.read(item_name))

                if not found_styles:
                    output_zip.writestr("xl/styles.xml", MINIMAL_STYLES_XML)

        return output_buffer.getvalue()
    except Exception:
        return None


def find_value_by_key(workbook, sheet_name, key, key_column, value_column):
    """
    Find a value in a workbook sheet by matching a key in a specific column.

    Args:
        workbook: OpenPyXL workbook object
        sheet_name (str): Sheet name to search in
        key (str): Key to search for
        key_column (int): 1-based column index of the key
        value_column (int): 1-based column index of the value to return

    Returns:
        The matched cell value, or 0 if not found.
    """
    resolved_sheet_name = _resolve_sheet_name(workbook, sheet_name)
    if resolved_sheet_name is None:
        return 0

    worksheet = workbook[resolved_sheet_name]
    normalized_key = _normalize_text(key)

    for row in worksheet.iter_rows():
        if len(row) < max(key_column, value_column):
            continue

        cell_value = _normalize_text(row[key_column - 1].value)
        if cell_value == normalized_key:
            return row[value_column - 1].value

    return 0


def fill_es_template_informa(
    template_bytes: bytes,
    registration_number: str,
    uploaded_files_by_name: Dict[str, bytes],
    source_key_column: int = 2,
    source_value_column: int = 4,
) -> bytes:
    """
    Fill the ES template in memory for Informa source files.

    Expected source file name pattern in uploads: <registration_number>_<year>.xlsx

    Returns:
        bytes: Filled template as xlsx bytes.
    """
    template_wb = _load_workbook_safe(template_bytes, data_only=False)
    template_ws = template_wb.active

    source_workbooks = {}

    try:
        for row in template_ws.iter_rows():
            if len(row) < 4:
                continue
            if row[0].value is None or row[1].value is None or row[3].value is None:
                continue

            source_sheet = _normalize_cell_value(row[0].value)
            key = _normalize_cell_value(row[1].value)
            year = _normalize_cell_value(row[3].value)
            source_file_name, source_file_payload = _get_uploaded_file_payload(
                uploaded_files_by_name,
                registration_number,
                year,
            )
            if source_file_payload is None:
                row[2].value = 0
                continue

            if source_file_name not in source_workbooks:
                try:
                    source_workbooks[source_file_name] = _load_workbook_safe(
                        source_file_payload,
                        data_only=True,
                    )
                except Exception as ex:
                    raise ValueError(f"Error reading source file '{source_file_name}': {ex}") from ex

            source_wb = source_workbooks[source_file_name]
            value = find_value_by_key(
                source_wb,
                source_sheet,
                key,
                source_key_column,
                source_value_column,
            )
            row[2].value = value

        output = io.BytesIO()
        template_wb.save(output)
        return output.getvalue()

    finally:
        for wb in source_workbooks.values():
            try:
                wb.close()
            except Exception:
                pass
        try:
            template_wb.close()
        except Exception:
            pass


def fill_fr_template_pappers(
    template_bytes: bytes,
    registration_number: str,
    uploaded_files_by_name: Dict[str, bytes],
) -> bytes:
    """
    Fill the FR template in memory for Pappers source files.

    Template structure (row 0 = header):
      col 0: sheet name
      col 1: key to look up
      col 2: description (ignored)
      col 3+: one column per year (header row contains the year value)

    Expected source file name pattern in uploads: <registration_number>_<year>.xlsx

        Pappers lookup mapping by sheet:
            - Actif: key col 9, value col 15
            - Passif: key col 17, value col 18
            - Compte de resultat: key col 12, value col 13
            - Compte de resultat (autre|suite): key col 14, value col 15

        Returns:
        bytes: Filled template as xlsx bytes.
    """
    template_wb = _load_workbook_safe(template_bytes, data_only=False)
    template_ws = template_wb.active

    source_workbooks = {}

    try:
        # Discover year columns from header row (row 1, 0-based col index → year string)
        header_row = next(template_ws.iter_rows(min_row=1, max_row=1, values_only=True))
        year_columns = {}  # 0-based col index -> year string
        for col_idx in range(3, len(header_row)):
            year_val = header_row[col_idx]
            if year_val is not None:
                year_columns[col_idx] = _normalize_cell_value(year_val)

        # Fill data rows (skip header row 1)
        for row in template_ws.iter_rows(min_row=2):
            if row[0].value is None or row[1].value is None:
                continue

            source_sheet = _normalize_cell_value(row[0].value)
            key = _normalize_cell_value(row[1].value)

            normalized_sheet = _normalize_text(source_sheet)
            if normalized_sheet == "actif":
                source_key_column = 9
                source_value_column = 15
            elif normalized_sheet == "passif":
                source_key_column = 17
                source_value_column = 18
            elif normalized_sheet == "compte de resultat":
                source_key_column = 12
                source_value_column = 13
            elif normalized_sheet in {
                "compte de resultat (autre)",
                "compte de resultat (suite)",
            }:
                source_key_column = 14
                source_value_column = 15
            else:
                # Unknown section in template; keep values as 0 for all years.
                for col_idx in year_columns:
                    row[col_idx].value = 0
                continue

            for col_idx, year in year_columns.items():
                source_file_name, source_file_payload = _get_uploaded_file_payload(
                    uploaded_files_by_name,
                    registration_number,
                    year,
                )
                if source_file_payload is None:
                    row[col_idx].value = 0
                    continue

                if source_file_name not in source_workbooks:
                    try:
                        source_workbooks[source_file_name] = _load_workbook_safe(
                            source_file_payload,
                            data_only=True,
                        )
                    except Exception as ex:
                        raise ValueError(f"Error reading source file '{source_file_name}': {ex}") from ex

                source_wb = source_workbooks[source_file_name]
                value = find_value_by_key(
                    source_wb,
                    source_sheet,
                    key,
                    source_key_column,
                    source_value_column,
                )
                row[col_idx].value = value

        output = io.BytesIO()
        template_wb.save(output)
        return output.getvalue()

    finally:
        for wb in source_workbooks.values():
            try:
                wb.close()
            except Exception:
                pass
        try:
            template_wb.close()
        except Exception:
            pass
